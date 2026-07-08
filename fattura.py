"""Lettura e validazione di un export fattura generato da Navision.

Un export valido è un file .xlsx che contiene almeno:
- il foglio "Generale" con i valori Nr., Cliente, Data documento;
- il foglio "Visualizzazione - Fatture vend" con le intestazioni attese in riga 2.

Dalle righe fattura si tengono solo:
- le righe con un valore nella colonna J (Importo riga IVA esclusa);
- le righe "DDT Nr. ..." (colonna D), che non hanno importo ma vanno nel file finale.
"""
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

FOGLIO_GENERALE = "Generale"
FOGLIO_RIGHE = "Visualizzazione - Fatture vend"

# Colonna (1 = A) -> testo atteso nella riga 2 del foglio righe.
# Se Navision cambia layout, la lettura si ferma qui con un errore esplicito.
INTESTAZIONI_ATTESE = {
    1: "Tipo",
    4: "Descrizione",
    6: "Quantità",
    7: "Cod. unità di misura",
    10: "Importo riga IVA esclusa",
}


class FatturaNonValida(ValueError):
    """Il file non è un export fattura utilizzabile. Il messaggio spiega il motivo."""


@dataclass(frozen=True)
class Riga:
    tipo: str            # colonna A ("" per le righe DDT)
    descrizione: str     # colonna D
    quantita: float | None  # colonna F
    unita_misura: str    # colonna G
    importo: float | None   # colonna J ("None" solo per le righe DDT)


@dataclass(frozen=True)
class Fattura:
    numero: str
    cliente: str
    data_documento: date
    righe: tuple[Riga, ...]


def leggi_fattura(percorso: Path) -> Fattura:
    """Legge un export di Navision e restituisce la Fattura. Solleva FatturaNonValida se il file non è conforme."""
    try:
        wb = load_workbook(percorso, read_only=True, data_only=True)
    except Exception as errore:
        raise FatturaNonValida(f"file non leggibile come xlsx: {errore}") from errore
    try:
        for foglio in (FOGLIO_GENERALE, FOGLIO_RIGHE):
            if foglio not in wb.sheetnames:
                raise FatturaNonValida(f"manca il foglio '{foglio}' (trovati: {', '.join(wb.sheetnames)})")
        numero, cliente, data_documento = _leggi_generale(wb[FOGLIO_GENERALE])
        righe = _leggi_righe(wb[FOGLIO_RIGHE])
    finally:
        wb.close()
    if not any(riga.importo is not None for riga in righe):
        raise FatturaNonValida("nessuna riga con importo nella colonna J")
    return Fattura(numero, cliente, data_documento, tuple(righe))


def _leggi_generale(ws) -> tuple[str, str, date]:
    """Estrae dal foglio 'Generale' (coppie chiave-valore in colonne A e B) i quattro campi richiesti."""
    valori: dict[str, object] = {}
    for riga in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        chiave = riga[0]
        valore = riga[1] if len(riga) > 1 else None
        if chiave is not None:
            valori[str(chiave).strip()] = valore

    richieste = ("Nr.", "Cliente", "Data documento", "Cod. agente")
    mancanti = [c for c in richieste if valori.get(c) is None or str(valori[c]).strip() == ""]
    if mancanti:
        raise FatturaNonValida(f"valori mancanti nel foglio '{FOGLIO_GENERALE}': {', '.join(mancanti)}")

    data = valori["Data documento"]
    if not isinstance(data, datetime):
        raise FatturaNonValida(f"'Data documento' non è una data: {data!r}")

    return (
        str(valori["Nr."]).strip(),
        str(valori["Cliente"]).strip(),
        data.date(),
    )


def _leggi_righe(ws) -> list[Riga]:
    """Legge le righe fattura, verificando prima le intestazioni in riga 2."""
    righe: list[Riga] = []
    intestazione_verificata = False
    for riga in ws.iter_rows(min_row=2, values_only=True):
        if len(riga) < 10:
            riga = riga + (None,) * (10 - len(riga))
        if not intestazione_verificata:
            for colonna, atteso in INTESTAZIONI_ATTESE.items():
                trovato = riga[colonna - 1]
                if trovato != atteso:
                    raise FatturaNonValida(
                        f"intestazione inattesa nel foglio righe, colonna {colonna}: "
                        f"{trovato!r} invece di {atteso!r}"
                    )
            intestazione_verificata = True
            continue

        tipo, descrizione, quantita, unita, importo = riga[0], riga[3], riga[5], riga[6], riga[9]
        testo = str(descrizione).strip() if descrizione is not None else ""
        if importo is not None:
            righe.append(Riga(str(tipo or "").strip(), testo, quantita, str(unita or "").strip(), importo))
        elif testo.startswith("DDT"):
            righe.append(Riga("", testo, None, "", None))

    if not intestazione_verificata:
        raise FatturaNonValida("il foglio delle righe è vuoto")
    return righe
