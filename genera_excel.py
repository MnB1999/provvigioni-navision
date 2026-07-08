"""Scrittura del file provvigioni: un unico file Excel per l'agente indicato.

Ogni file contiene due parti:

1. Il listato delle fatture (replica il formato del file usato finora):
   per ogni cliente in ordine alfabetico, le sue fatture in ordine di numero;
   la prima con intestazione "CLIENTE - Fattura nr NUMERO", le successive
   "ft NUMERO"; righe su 5 colonne (Tipo, Descrizione, Quantità, Unità,
   Importo) con una riga vuota tra i blocchi.

2. La tabella di calcolo provvigionale, sette righe più in basso:
   una riga per voce con formule =SUM(...) che sommano gli importi del
   listato per tipo di voce, percentuale a 0 da inserire a mano,
   Provvigione = Importo x percentuale e riga del totale.
   "Materiali" è sempre a 0: si compila a mano.
"""
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from classificazione import VOCI, VoceNonClassificabile, voce_per
from fattura import Fattura

CARATTERE_TITOLO = Font(name="Arial", size=12, bold=True)
CARATTERE_DATI = Font(name="Verdana", size=10)
CARATTERE_TABELLA = Font(name="Arial", size=10)
CARATTERE_TABELLA_INTESTAZIONE = Font(name="Arial", size=9, bold=True)
BORDO = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))
FORMATO_EURO = '#,##0.00\\ "€"'
FORMATO_PERCENTO = "0.00%"
LARGHEZZE_COLONNE = {"A": 12, "B": 52, "C": 12, "D": 12, "E": 13}
RIGHE_VUOTE_PRIMA_DELLA_TABELLA = 6
LIMITE_ARGOMENTI_SUM = 255  # limite di Excel


def scrivi_provvigioni(fatture: list[Fattura], cartella: Path, agente: str) -> Path:
    """Scrive un unico file provvigioni con tutte le fatture, intitolato all'agente indicato. Restituisce il percorso creato."""
    if not fatture:
        raise ValueError("nessuna fattura da elaborare")
    if not agente.strip():
        raise ValueError("il nome dell'agente è vuoto")
    _verifica_classificazione(fatture)
    cartella.mkdir(parents=True, exist_ok=True)
    return _scrivi_file_agente(agente.strip(), fatture, cartella)


def _verifica_classificazione(fatture: list[Fattura]) -> None:
    """Prima di scrivere qualsiasi file, tutte le righe con importo devono essere attribuibili a una voce."""
    sconosciute = set()
    for fattura in fatture:
        for riga in fattura.righe:
            if riga.importo is not None:
                try:
                    voce_per(riga.descrizione)
                except VoceNonClassificabile:
                    sconosciute.add(riga.descrizione)
    if sconosciute:
        elenco = "\n  ".join(sorted(sconosciute))
        raise VoceNonClassificabile(
            f"descrizioni non attribuibili ad alcuna voce della tabella:\n  {elenco}\n"
            "Aggiungere le regole corrispondenti in classificazione.py."
        )


def _scrivi_file_agente(agente: str, fatture: list[Fattura], cartella: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[\[\]:*?/\\]", "_", agente)[:31]
    for colonna, larghezza in LARGHEZZE_COLONNE.items():
        ws.column_dimensions[colonna].width = larghezza

    righe_per_voce = _scrivi_listato(ws, fatture)
    _scrivi_tabella_calcolo(ws, righe_per_voce)

    nome_sicuro = re.sub(r"[^\w-]", "_", agente)
    percorso = cartella / f"Provvigioni_{nome_sicuro}.xlsx"
    wb.save(percorso)
    return percorso


def _scrivi_listato(ws, fatture: list[Fattura]) -> dict[str, list[int]]:
    """Scrive il listato delle fatture e restituisce, per ogni voce, i numeri
    di riga Excel delle righe con importo che le appartengono."""
    righe_per_voce: dict[str, list[int]] = defaultdict(list)

    per_cliente: dict[str, list[Fattura]] = defaultdict(list)
    for fattura in fatture:
        per_cliente[fattura.cliente].append(fattura)

    for cliente in sorted(per_cliente):
        in_ordine = sorted(per_cliente[cliente], key=lambda f: f.numero)
        for posizione, fattura in enumerate(in_ordine):
            if posizione == 0:
                titolo = f"{cliente} - Fattura nr {fattura.numero}"
            else:
                titolo = f"ft {fattura.numero}"
            ws.append([titolo])
            ws.cell(row=ws.max_row, column=1).font = CARATTERE_TITOLO
            for riga in fattura.righe:
                ws.append([riga.tipo, riga.descrizione, riga.quantita, riga.unita_misura, riga.importo])
                for cella in ws[ws.max_row]:
                    cella.font = CARATTERE_DATI
                if riga.importo is not None:
                    righe_per_voce[voce_per(riga.descrizione)].append(ws.max_row)
            ws.append([None])  # riga vuota di separazione

    return righe_per_voce


def _scrivi_tabella_calcolo(ws, righe_per_voce: dict[str, list[int]]) -> None:
    # max_row include già la riga vuota di separazione finale del listato,
    # che conta come prima delle sei righe vuote prima della tabella.
    riga = ws.max_row + RIGHE_VUOTE_PRIMA_DELLA_TABELLA

    intestazioni = [("B", "Descrizione"), ("C", "Provvigione %\nsu fatturato"), ("D", "Importo"), ("E", "Provvigione")]
    for colonna, testo in intestazioni:
        cella = ws[f"{colonna}{riga}"]
        cella.value = testo
        cella.font = CARATTERE_TABELLA_INTESTAZIONE
        cella.border = BORDO
        cella.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[riga].height = 26

    prima_voce = riga + 1
    for voce in VOCI:
        riga += 1
        righe = righe_per_voce.get(voce, [])
        importo = _formula_somma(righe) if righe else 0  # "Materiali" non è mai in righe_per_voce: resta 0
        valori = [("B", voce, "General"), ("C", 0, FORMATO_PERCENTO),
                  ("D", importo, FORMATO_EURO), ("E", f"=D{riga}*C{riga}", FORMATO_EURO)]
        for colonna, valore, formato in valori:
            cella = ws[f"{colonna}{riga}"]
            cella.value = valore
            cella.number_format = formato
            cella.font = CARATTERE_TABELLA
            cella.border = BORDO

    riga += 1
    ws[f"D{riga}"].value = "totale:"
    ws[f"D{riga}"].font = CARATTERE_TABELLA
    ws[f"E{riga}"].value = f"=SUM(E{prima_voce}:E{riga - 1})"
    ws[f"E{riga}"].number_format = FORMATO_EURO
    ws[f"E{riga}"].font = CARATTERE_TABELLA


def _formula_somma(righe_excel: list[int]) -> str:
    """Costruisce =SUM(...) sugli importi in colonna E, compattando le righe consecutive in intervalli."""
    intervalli: list[tuple[int, int]] = []
    inizio = fine = righe_excel[0]
    for numero in righe_excel[1:]:
        if numero == fine + 1:
            fine = numero
        else:
            intervalli.append((inizio, fine))
            inizio = fine = numero
    intervalli.append((inizio, fine))

    argomenti = [f"E{a}" if a == b else f"E{a}:E{b}" for a, b in intervalli]
    if len(argomenti) > LIMITE_ARGOMENTI_SUM:
        raise ValueError(f"la formula SUM avrebbe {len(argomenti)} argomenti: Excel ne accetta al massimo {LIMITE_ARGOMENTI_SUM}")
    return "=SUM(" + ",".join(argomenti) + ")"
