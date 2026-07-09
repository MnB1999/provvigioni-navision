"""Test della pipeline senza interfaccia grafica.

Crea export finti con la stessa struttura di quelli di Navision
e verifica lettura, validazione, classificazione e scrittura dei file
provvigioniali con la tabella di calcolo.

Esecuzione:  py test_fattura.py
"""
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from classificazione import VoceNonClassificabile, voce_per
from fattura import FatturaNonValida, leggi_fattura
from genera_excel import scrivi_provvigioni

INTESTAZIONI = [
    "Tipo", "Nr.", "Includi in report transazioni IVA", "Descrizione",
    "Descrizione Corrente", "Quantità", "Cod. unità di misura",
    "Prezzo unitario IVA esclusa", "% sconto riga", "Importo riga IVA esclusa",
]

RIGA_DDT = [" ", None, "No", "DDT Nr. TEST-01 - 12/02/26:", None, None, "", None, None, None]
RIGA_ELIO = ["Articolo", "X1", "No", "ELIO 5.0 Bombola MEDIA", "ELIO 5.0 Bombola MEDIA", 2, "BLA14", 100, 0, 200]
RIGA_NOLO = ["Risorsa", "X2", "No", "Nolo Mensile Bombola", "Nolo Mensile Bombola", 3, "PZ", 5, 0, 15]
RIGA_SENZA_IMPORTO = ["Risorsa", "X3", "No", "Omaggio sconto 100%", "Omaggio", 1, "PZ", 10, 100, None]
RIGA_IGNOTA = ["Articolo", "X4", "No", "PRODOTTO MISTERIOSO", "PRODOTTO MISTERIOSO", 1, "PZ", 5, 0, 5]


def crea_export(percorso: Path, numero: str, cliente: str, agente: str,
                righe: list, data: datetime = datetime(2026, 2, 12),
                senza_foglio_righe: bool = False, intestazioni: list = INTESTAZIONI) -> None:
    wb = Workbook()
    generale = wb.active
    generale.title = "Generale"
    generale.append([f"Visualizzazione - Fatture vendita reg. - {numero}"])
    for chiave, valore in [("Nr.", numero), ("Cliente", cliente),
                           ("Data documento", data), ("Cod. agente", agente)]:
        generale.append([chiave, valore])
    if not senza_foglio_righe:
        foglio = wb.create_sheet("Visualizzazione - Fatture vend")
        foglio.append([f"Visualizzazione - Fatture vendita reg. - {numero}"])
        foglio.append(intestazioni)
        for riga in righe:
            foglio.append(riga)
    wb.save(percorso)


def test_lettura(tmp: Path) -> None:
    percorso = tmp / "f1.xlsx"
    crea_export(percorso, "26/000001", "CLIENTE UNO ", "GCS",
                [RIGA_DDT, RIGA_ELIO, RIGA_SENZA_IMPORTO, RIGA_NOLO])
    f = leggi_fattura(percorso)
    assert f.numero == "26/000001"
    assert f.cliente == "CLIENTE UNO"  # spazio finale rimosso
    assert f.data_documento.isoformat() == "2026-02-12"
    # tenute: DDT + 2 righe con importo; esclusa quella con J vuota
    assert len(f.righe) == 3
    assert f.righe[0].descrizione.startswith("DDT") and f.righe[0].importo is None
    assert f.righe[1].importo == 200 and f.righe[1].quantita == 2 and f.righe[1].unita_misura == "BLA14"
    assert f.righe[2].importo == 15


def _si_aspetta_errore(percorso: Path, frammento: str) -> None:
    try:
        leggi_fattura(percorso)
    except FatturaNonValida as errore:
        assert frammento in str(errore), f"messaggio inatteso: {errore}"
    else:
        raise AssertionError(f"atteso errore contenente {frammento!r}")


def test_errori(tmp: Path) -> None:
    p1 = tmp / "senza_foglio.xlsx"
    crea_export(p1, "26/000002", "CLIENTE", "GCS", [RIGA_ELIO], senza_foglio_righe=True)
    _si_aspetta_errore(p1, "manca il foglio")

    p2 = tmp / "intestazioni_sbagliate.xlsx"
    diverse = INTESTAZIONI.copy()
    diverse[9] = "Importo totale"
    crea_export(p2, "26/000003", "CLIENTE", "GCS", [RIGA_ELIO], intestazioni=diverse)
    _si_aspetta_errore(p2, "intestazione inattesa")

    p3 = tmp / "senza_importi.xlsx"
    crea_export(p3, "26/000004", "CLIENTE", "GCS", [RIGA_DDT, RIGA_SENZA_IMPORTO])
    _si_aspetta_errore(p3, "nessuna riga con importo")

    p4 = tmp / "senza_agente.xlsx"
    crea_export(p4, "26/000005", "CLIENTE", "", [RIGA_ELIO])
    assert leggi_fattura(p4).numero == "26/000005"

    p5 = tmp / "non_excel.xlsx"
    p5.write_text("questo non è un xlsx")
    _si_aspetta_errore(p5, "non leggibile")


CLASSIFICAZIONI_DI_RIFERIMENTO = {
    "Adr": ["CONTRIB. ADR BB", "CONTRIBUTO ADR BB", "CONTRIBUTO ADR PACCO"],
    "Aqs e addizionali vari": [
        "ADDIZIONALE TRASPORTO BB IN PK / BB IN PK 300", "ADDIZIONALE TRASPORTO BOMBOLA",
        "ADEMPIMENTI QUAL/SICUR", "CONTRIBUTO ENERGIA BB / BB IN PK",
    ],
    "Elio forza maggiore": ["Elio Forza Maggiore"],
    "Gas alimentari e puri": [
        "TRESARIS C Pacco", "TRESARIS N Pacco",
        "TRESARIS NC30 Bombola GRANDE", "TRESARIS OCN 25-10 Bombola GRANDE",
    ],
    "Gas refrigeranti, smaltimento refrigeranti e propano": [
        "GAS REFRIGERANTE R32 Bombola PICCOLA", "GAS REFRIGERANTE R410A Bombola GRANDE",
        "GAS REFRIGERANTE R410A Bombola PICCOLA", "GAS REFRIGERANTE R452A (XP44) Bombola PICCOLA",
        "GAS REFRIGERANTE R452A -XP44- Bombola GRANDE", "PROPANO Bombola PICCOLA",
    ],
    "Gas tecnici": [
        "ACETILENE TECNICO Bombola GRANDE", "ACETILENE TECNICO Bombola MEDIA", "ACETILENE TECNICO Bombola PICCOLA",
        "AZOTO 5.0 Bombola GRANDE", "AZOTO 5.0 Bombola MEDIA", "AZOTO 5.0 Bombola PICCOLA",
        "CORGON 18 - Competence Bombola PICCOLA",
        "OSSIGENO TECNICO Bombola GRANDE", "OSSIGENO TECNICO Bombola MEDIA", "OSSIGENO TECNICO Bombola PICCOLA",
        "ELIO 5.0 Bombola GRANDE", "ELIO 5.0 Bombola MEDIA",
    ],
    "Nolo e RMR": [
        "Conguaglio 2025 nolo giornal. bombole Gas Speciali", "Nolo Giornaliero Bombola",
        "Nolo Giornaliero Bombola - GAS SPECIALI", "Nolo Mensile Bombola",
        "Nolo Mensile Bombola - GAS SPECIALI", "Nolo Mensile Bombola Grande - Gas Refr.",
        "Nolo Mensile Bombola Piccola - Gas Refr.", "Nolo Mensile PACCO",
        "RIMBORSO MANCATA RESA Bombola Grande - Gas Refr.",
        "RIMBORSO MANCATA RESA Bombola Piccola - Gar Refr.", "RMR Bombola",
    ],
    "Trasporto": ["TRASPORTO BOMBOLA", "TRASPORTO PACCO", "TRASPORTO PER VIAGGIO N/A"],
}


def test_classificazione() -> None:
    for voce, descrizioni in CLASSIFICAZIONI_DI_RIFERIMENTO.items():
        for descrizione in descrizioni:
            trovata = voce_per(descrizione)
            assert trovata == voce, f"{descrizione!r}: attesa {voce!r}, trovata {trovata!r}"
    # gradi di purezza: qualunque gas con 5.0/5.5/6.0 è "puro", salvo AZOTO 5.0 e ELIO 5.0
    assert voce_per("ARGON 6.0 Bombola PICCOLA") == "Gas alimentari e puri"
    assert voce_per("IDROGENO 5.5 Bombola GRANDE") == "Gas alimentari e puri"
    assert voce_per("OSSIGENO 6.0 Bombola MEDIA") == "Gas alimentari e puri"
    assert voce_per("AZOTO 5.0 Bombola MEDIA") == "Gas tecnici"
    assert voce_per("ELIO 5.0 Bombola GRANDE") == "Gas tecnici"
    try:
        voce_per("PRODOTTO MISTERIOSO")
    except VoceNonClassificabile:
        pass
    else:
        raise AssertionError("attesa VoceNonClassificabile")


def test_scrittura(tmp: Path) -> None:
    sorgenti = tmp / "sorgenti"
    sorgenti.mkdir()
    dati = [
        # (numero, cliente, righe) — ordine di arrivo volutamente mescolato
        ("26/000030", "BETA SRL", [RIGA_NOLO]),
        ("26/000010", "ALFA SPA", [RIGA_DDT, RIGA_ELIO]),
        ("26/000020", "ALFA SPA", [RIGA_NOLO]),
    ]
    fatture = []
    for indice, (numero, cliente, righe) in enumerate(dati):
        percorso = sorgenti / f"f{indice}.xlsx"
        crea_export(percorso, numero, cliente, "GCS", righe)
        fatture.append(leggi_fattura(percorso))

    uscita = tmp / "output"
    creato = scrivi_provvigioni(fatture, uscita, "GCS")
    assert creato.name == "Provvigioni_GCS.xlsx", creato.name

    # il nome agente è quello inserito dall'operatore, non il codice rilevato
    try:
        scrivi_provvigioni(fatture, uscita, "   ")
    except ValueError:
        pass
    else:
        raise AssertionError("atteso ValueError per nome agente vuoto")

    ws = load_workbook(uscita / "Provvigioni_GCS.xlsx")["GCS"]

    listato = [[c.value for c in riga[:5]] for riga in ws.iter_rows(min_row=1, max_row=9)]
    atteso = [
        ["ALFA SPA - Fattura nr 26/000010", None, None, None, None],
        [None, "DDT Nr. TEST-01 - 12/02/26:", None, None, None],
        ["Articolo", "ELIO 5.0 Bombola MEDIA", 2, "BLA14", 200],
        [None, None, None, None, None],
        ["ft 26/000020", None, None, None, None],
        ["Risorsa", "Nolo Mensile Bombola", 3, "PZ", 15],
        [None, None, None, None, None],
        ["BETA SRL - Fattura nr 26/000030", None, None, None, None],
        ["Risorsa", "Nolo Mensile Bombola", 3, "PZ", 15],
    ]
    assert listato == atteso, f"listato diverso dall'atteso:\n{listato}"
    assert ws["A1"].font.bold and ws["A1"].font.name == "Arial"
    assert not ws["A3"].font.bold and ws["A3"].font.name == "Verdana"

    # tabella di calcolo
    assert ws["B16"].value == "Descrizione" and ws["B16"].font.bold
    assert ws["D16"].value == "Importo" and ws["E16"].value == "Provvigione"
    tabella = {ws[f"B{r}"].value: (ws[f"C{r}"].value, ws[f"D{r}"].value, ws[f"E{r}"].value)
               for r in range(17, 26)}
    attesa = {
        "Gas tecnici": (0, "=SUM(E3)", "=D17*C17"),
        "Elio forza maggiore": (0, 0, "=D18*C18"),
        "Gas refrigeranti, smaltimento refrigeranti e propano": (0, 0, "=D19*C19"),
        "Gas alimentari e puri": (0, 0, "=D20*C20"),
        "Adr": (0, 0, "=D21*C21"),
        "Trasporto": (0, 0, "=D22*C22"),
        "Aqs e addizionali vari": (0, 0, "=D23*C23"),
        "Materiali": (0, 0, "=D24*C24"),
        "Nolo e RMR": (0, "=SUM(E6,E9)", "=D25*C25"),
    }
    assert tabella == attesa, f"tabella diversa dall'attesa:\n{tabella}"
    assert ws["C17"].number_format == "0.00%"
    assert ws["D17"].number_format == '#,##0.00\\ "€"'
    assert ws["D26"].value == "totale:" and ws["E26"].value == "=SUM(E17:E25)"


def test_intervalli_consecutivi(tmp: Path) -> None:
    percorso = tmp / "consecutive.xlsx"
    crea_export(percorso, "26/000050", "DELTA SRL", "GCS", [RIGA_NOLO, RIGA_NOLO, RIGA_NOLO])
    uscita = tmp / "output_intervalli"
    scrivi_provvigioni([leggi_fattura(percorso)], uscita, "GCS")
    ws = load_workbook(uscita / "Provvigioni_GCS.xlsx")["GCS"]
    formule = {ws[f"B{r}"].value: ws[f"D{r}"].value for r in range(12, 21)}
    assert formule["Nolo e RMR"] == "=SUM(E2:E4)", formule["Nolo e RMR"]


def test_descrizione_sconosciuta(tmp: Path) -> None:
    percorso = tmp / "ignota.xlsx"
    crea_export(percorso, "26/000060", "EPSILON SRL", "GCS", [RIGA_IGNOTA])
    fattura = leggi_fattura(percorso)
    try:
        scrivi_provvigioni([fattura], tmp / "output_ignota", "GCS")
    except VoceNonClassificabile as errore:
        assert "PRODOTTO MISTERIOSO" in str(errore)
        assert "classificazione.py" in str(errore)
    else:
        raise AssertionError("attesa VoceNonClassificabile")


def main() -> None:
    with tempfile.TemporaryDirectory() as cartella:
        tmp = Path(cartella)
        test_lettura(tmp)
        test_errori(tmp)
        test_classificazione()
        test_scrittura(tmp)
        test_intervalli_consecutivi(tmp)
        test_descrizione_sconosciuta(tmp)
    print("Tutti i test superati.")


if __name__ == "__main__":
    main()
